"""
Export — CSV / Excel / PDF pour contacts et ROC.
"""

import csv
import io
from datetime import datetime
from flask import Blueprint, request, send_file, jsonify
from flask_login import login_required, current_user
from reportlab.lib.pagesizes import A3, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_LEFT, TA_CENTER
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app import db
from app.models import Contact, Roc, AuditLog

export_bp = Blueprint("export", __name__)


def _log(action: str, detail: str = "") -> None:
    entry = AuditLog(
        username=current_user.username,
        action="EXPORT",
        table_name=action,
        detail=detail,
        ip_address=request.remote_addr or "",
    )
    db.session.add(entry)
    db.session.commit()


def _get_data(table: str) -> tuple[list, list]:
    """Retourne (headers, rows) selon la table demandée."""
    if table == "contacts":
        headers = ["Société", "Nom", "Prénom", "Fonction", "Email",
                   "Téléphone", "Téléphone 2", "Notes"]
        rows = [
            [c.societe, c.nom, c.prenom, c.fonction,
             c.email, c.telephone, c.telephone2, c.notes]
            for c in Contact.query.order_by(Contact.societe).all()
        ]
    else:
        headers = ["Nom Client", "ROC", "Trinity", "Infogérance",
                   "Astreinte", "Type Contrat", "Date Anniversaire Contrat"]
        rows = [
            [r.nom_client, r.roc, r.trinity, r.infogerance,
             r.astreinte, r.type_contrat, r.date_anniversaire_contrat]
            for r in Roc.query.order_by(Roc.nom_client).all()
        ]
    return headers, rows


# ── CSV ───────────────────────────────────────────────────────────────────────

@export_bp.route("/<table>/csv")
@login_required
def export_csv(table: str):
    if table not in ("contacts", "rocs"):
        return jsonify({"error": "Table invalide."}), 400

    headers, rows = _get_data(table)
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(headers)
    writer.writerows(rows)
    buf.seek(0)

    _log(table, f"CSV — {len(rows)} lignes")
    fname = f"annuaire_{table}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return send_file(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=fname,
    )


# ── Excel ─────────────────────────────────────────────────────────────────────

@export_bp.route("/<table>/xlsx")
@login_required
def export_xlsx(table: str):
    if table not in ("contacts", "rocs"):
        return jsonify({"error": "Table invalide."}), 400

    headers, rows = _get_data(table)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table.capitalize()

    # En-tête
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Données
    alt_fill = PatternFill("solid", fgColor="EEF2FF")
    for row_idx, row in enumerate(rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val or "")
            if fill:
                cell.fill = fill

    # Largeurs auto
    col_widths = [22, 15, 15, 20, 30, 16, 16, 40] if table == "contacts" \
        else [25, 12, 12, 20, 20, 18, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    _log(table, f"Excel — {len(rows)} lignes")
    fname = f"annuaire_{table}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


# ── PDF ───────────────────────────────────────────────────────────────────────

@export_bp.route("/<table>/pdf")
@login_required
def export_pdf(table: str):
    if table not in ("contacts", "rocs"):
        return jsonify({"error": "Table invalide."}), 400

    headers, rows = _get_data(table)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A3),
        leftMargin=12*mm, rightMargin=12*mm,
        topMargin=12*mm, bottomMargin=12*mm,
    )

    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("cell", parent=styles["Normal"],
                                fontSize=7, leading=9)
    head_style = ParagraphStyle("head", parent=styles["Normal"],
                                fontSize=8, leading=10,
                                textColor=colors.white, fontName="Helvetica-Bold")

    usable = landscape(A3)[0] - 24*mm
    weights = [3, 2, 2, 2.5, 4, 2, 2, 4] if table == "contacts" \
        else [3.5, 2, 2, 2.5, 2.5, 2, 3]
    total_w = sum(weights)
    col_widths = [usable * (w / total_w) for w in weights]

    # Titre
    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                 fontSize=14, spaceAfter=4)
    sub_style   = ParagraphStyle("sub", parent=styles["Normal"],
                                 fontSize=9, textColor=colors.grey, spaceAfter=8)

    table_label = "Contacts" if table == "contacts" else "ROC"
    story = [
        Paragraph(f"Annuaire Neoedge — {table_label}", title_style),
        Paragraph(
            f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')} "
            f"par {current_user.full_name or current_user.username} — "
            f"{len(rows)} enregistrement(s)",
            sub_style,
        ),
        Spacer(1, 4),
    ]

    # Tableau
    head_row = [Paragraph(h, head_style) for h in headers]
    data_rows = [head_row] + [
        [Paragraph(str(v or ""), cell_style) for v in row]
        for row in rows
    ]

    tbl = Table(data_rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 8),
        ("FONTSIZE",    (0, 1), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#EEF2FF")]),
        ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN",      (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING",  (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",(0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)
    doc.build(story)
    buf.seek(0)

    _log(table, f"PDF — {len(rows)} lignes")
    fname = f"annuaire_{table}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name=fname)


# ══════════════════════════════════════════════════════════════════════════════
# IMPORT CSV
# ══════════════════════════════════════════════════════════════════════════════

# Correspondance entre noms de colonnes CSV possibles → champ modèle
_CONTACT_ALIASES = {
    "societe":    ["societe","société","company","entreprise","organisation","organization"],
    "nom":        ["nom","lastname","last_name","name"],
    "prenom":     ["prenom","prénom","firstname","first_name","given_name"],
    "fonction":   ["fonction","function","poste","job","title","titre"],
    "email":      ["email","mail","e-mail","courriel"],
    "telephone":  ["telephone","téléphone","tel","phone","mobile","portable"],
    "telephone2": ["telephone2","tel2","phone2","mobile2","portable2","telephone_2"],
    "notes":      ["notes","note","commentaire","commentaires","remarks","comment"],
}
_ROC_ALIASES = {
    "nom_client":                ["nom_client","client","nom client","customer","name"],
    "roc":                       ["roc"],
    "trinity":                   ["trinity"],
    "infogerance":               ["infogerance","infogérance","managed","gestion"],
    "astreinte":                 ["astreinte","on_call","oncall"],
    "type_contrat":              ["type_contrat","type contrat","contrat","contract","contract_type"],
    "date_anniversaire_contrat": ["date_anniversaire_contrat","date anniversaire","anniversaire","renewal","renouvellement"],
}


def _detect_delimiter(sample: str) -> str:
    """Détecte automatiquement le délimiteur (;  ,  |  tab)."""
    counts = {d: sample.count(d) for d in [";", ",", "|", "\t"]}
    return max(counts, key=counts.get)


def _map_headers(raw_headers: list[str], aliases: dict) -> dict:
    """Retourne {index_colonne: champ_modele} selon les aliases."""
    mapping = {}
    for idx, h in enumerate(raw_headers):
        h_clean = h.strip().lower().replace(" ", "_").replace("-", "_")
        for field, alts in aliases.items():
            if h_clean in alts and field not in mapping.values():
                mapping[idx] = field
                break
    return mapping


@export_bp.route("/import/csv", methods=["POST"])
@login_required
def import_csv():
    from flask_login import current_user
    if not current_user.can_write:
        return jsonify({"error": "Accès refusé (rôle editor ou admin requis)."}), 403

    table   = (request.form.get("table") or "").strip()
    mode    = (request.form.get("mode")  or "append").strip()   # append | replace
    if table not in ("contacts", "rocs"):
        return jsonify({"error": "Table invalide (contacts ou rocs)."}), 400
    if "file" not in request.files:
        return jsonify({"error": "Aucun fichier reçu."}), 400

    f = request.files["file"]
    try:
        raw = f.read().decode("utf-8-sig")   # gère le BOM Windows
    except UnicodeDecodeError:
        try:
            f.seek(0)
            raw = f.read().decode("latin-1")
        except Exception:
            return jsonify({"error": "Impossible de lire le fichier (encodage non supporté)."}), 400

    if not raw.strip():
        return jsonify({"error": "Fichier vide."}), 400

    delimiter = _detect_delimiter(raw[:2000])
    reader    = csv.DictReader(io.StringIO(raw), delimiter=delimiter)
    raw_headers = reader.fieldnames or []

    aliases = _CONTACT_ALIASES if table == "contacts" else _ROC_ALIASES
    mapping = _map_headers(list(raw_headers), aliases)

    if not mapping:
        return jsonify({
            "error": f"Aucune colonne reconnue. Colonnes trouvées : {', '.join(raw_headers)}. "
                     f"Colonnes attendues : {', '.join(aliases.keys())}."
        }), 400

    rows_raw = list(reader)
    if not rows_raw:
        return jsonify({"error": "Aucune ligne de données dans le fichier."}), 400

    inserted = 0
    skipped  = 0
    errors   = []

    try:
        if mode == "replace":
            if table == "contacts":
                Contact.query.delete(synchronize_session=False)
            else:
                Roc.query.delete(synchronize_session=False)
            db.session.flush()

        for line_no, row in enumerate(rows_raw, start=2):
            row_vals = {mapping[i]: v.strip() for i, v in enumerate(row.values()) if i in mapping}

            if table == "contacts":
                if not row_vals.get("societe") and not row_vals.get("nom") and not row_vals.get("email"):
                    skipped += 1
                    continue
                db.session.add(Contact(
                    societe    = row_vals.get("societe",    ""),
                    nom        = row_vals.get("nom",        ""),
                    prenom     = row_vals.get("prenom",     ""),
                    fonction   = row_vals.get("fonction",   ""),
                    email      = row_vals.get("email",      ""),
                    telephone  = row_vals.get("telephone",  ""),
                    telephone2 = row_vals.get("telephone2", ""),
                    notes      = row_vals.get("notes",      ""),
                    updated_by = current_user.username,
                ))
            else:
                if not row_vals.get("nom_client") and not row_vals.get("roc"):
                    skipped += 1
                    continue
                db.session.add(Roc(
                    nom_client                = row_vals.get("nom_client",               ""),
                    roc                       = row_vals.get("roc",                      ""),
                    trinity                   = row_vals.get("trinity",                  ""),
                    infogerance               = row_vals.get("infogerance",              ""),
                    astreinte                 = row_vals.get("astreinte",                ""),
                    type_contrat              = row_vals.get("type_contrat",             ""),
                    date_anniversaire_contrat = row_vals.get("date_anniversaire_contrat",""),
                    updated_by                = current_user.username,
                ))
            inserted += 1

        db.session.add(AuditLog(
            username   = current_user.username,
            action     = "IMPORT_CSV",
            table_name = table,
            detail     = f"Import CSV {table} — {inserted} insérés, {skipped} ignorés (mode: {mode}, délimiteur: '{delimiter}')",
            ip_address = request.remote_addr or "",
        ))
        db.session.commit()

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Erreur lors de l'insertion : {str(e)}"}), 500

    return jsonify({
        "ok":       True,
        "inserted": inserted,
        "skipped":  skipped,
        "mode":     mode,
        "delimiter": delimiter,
        "table":    table,
    })


# ── Aperçu CSV (preview 5 premières lignes avant import) ─────────────────────

@export_bp.route("/import/csv/preview", methods=["POST"])
@login_required
def preview_csv():
    if not current_user.can_write:
        return jsonify({"error": "Accès refusé."}), 403

    table = (request.form.get("table") or "contacts").strip()
    if "file" not in request.files:
        return jsonify({"error": "Aucun fichier."}), 400

    f = request.files["file"]
    try:
        raw = f.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        f.seek(0)
        raw = f.read().decode("latin-1")

    delimiter   = _detect_delimiter(raw[:2000])
    reader      = csv.DictReader(io.StringIO(raw), delimiter=delimiter)
    raw_headers = list(reader.fieldnames or [])
    aliases     = _CONTACT_ALIASES if table == "contacts" else _ROC_ALIASES
    mapping     = _map_headers(raw_headers, aliases)

    rows_raw = list(reader)
    preview  = []
    for row in rows_raw[:5]:
        vals = list(row.values())
        preview.append({mapping.get(i, f"col_{i}"): v.strip() for i, v in enumerate(vals)})

    return jsonify({
        "delimiter":       delimiter,
        "raw_headers":     raw_headers,
        "mapped_fields":   {str(i): f for i, f in mapping.items()},
        "total_rows":      len(rows_raw),
        "preview":         preview,
        "recognized_cols": list(mapping.values()),
        "missing_cols":    [f for f in aliases if f not in mapping.values()],
    })


# ══════════════════════════════════════════════════════════════════════════════
# IMPORT PDF
# ══════════════════════════════════════════════════════════════════════════════

def _extract_pdf_text(file_bytes: bytes) -> list[str]:
    """Extrait le texte d'un PDF page par page avec pypdf."""
    from pypdf import PdfReader
    import io as _io
    reader = PdfReader(_io.BytesIO(file_bytes))
    pages = []
    for page in reader.pages:
        text = page.extract_text() or ""
        if text.strip():
            pages.append(text)
    return pages


def _parse_pdf_contacts(pages: list[str]) -> list[dict]:
    """
    Tente de détecter des blocs contact dans le texte extrait du PDF.
    Stratégies : lignes avec @, lignes avec numéro de téléphone, blocs séparés.
    """
    import re
    contacts = []
    email_re = re.compile(r"[\w.+-]+@[\w.-]+\.\w{2,}")
    phone_re = re.compile(r"(?:\+33|0)[1-9][\s.\-]?(?:\d{2}[\s.\-]?){3}\d{2}")

    full_text = "\n".join(pages)
    # Séparer par double saut de ligne (blocs)
    blocks = re.split(r"\n{2,}", full_text.strip())

    for block in blocks:
        lines = [l.strip() for l in block.splitlines() if l.strip()]
        if not lines:
            continue

        contact = {}
        for line in lines:
            email_m = email_re.search(line)
            phone_m = phone_re.search(line)
            if email_m and not contact.get("email"):
                contact["email"] = email_m.group()
            elif phone_m and not contact.get("telephone"):
                contact["telephone"] = phone_m.group().replace(" ", "").replace(".", "").replace("-", "")
            elif not contact.get("societe") and len(line) > 2 and not email_m and not phone_m:
                # Heuristique : première ligne non-email/tel = société ou nom
                if not contact.get("nom"):
                    # Si la ligne contient un espace et ressemble à un nom (2 mots)
                    parts = line.split()
                    if len(parts) == 2 and all(p[0].isupper() for p in parts if p):
                        contact["nom"]    = parts[0]
                        contact["prenom"] = parts[1]
                    else:
                        contact["societe"] = line
                else:
                    contact.setdefault("fonction", line)

        # Ne garder que les blocs avec au moins email ou téléphone
        if contact.get("email") or contact.get("telephone") or contact.get("societe"):
            contacts.append(contact)

    return contacts


def _parse_pdf_table(pages: list[str], table: str) -> list[dict]:
    """
    Détecte une structure tabulaire dans le PDF (lignes avec séparateurs réguliers).
    Retourne une liste de dicts mappés sur les champs du modèle.
    """
    import re
    full_text = "\n".join(pages)
    lines = [l for l in full_text.splitlines() if l.strip()]

    # Chercher une ligne d'en-tête (contient des mots-clés connus)
    aliases = _CONTACT_ALIASES if table == "contacts" else _ROC_ALIASES
    header_idx = None
    detected_mapping = {}

    for i, line in enumerate(lines[:30]):   # en-tête dans les 30 premières lignes
        cells = re.split(r"\s{2,}|\t", line)  # séparer par 2+ espaces ou tab
        if len(cells) < 2:
            continue
        mapping = _map_headers(cells, aliases)
        if len(mapping) >= 2:
            header_idx    = i
            detected_mapping = mapping
            break

    if header_idx is None:
        return []

    rows = []
    for line in lines[header_idx + 1:]:
        cells = re.split(r"\s{2,}|\t", line)
        if len(cells) < 2:
            continue
        row = {detected_mapping[i]: v.strip() for i, v in enumerate(cells) if i in detected_mapping}
        if any(v for v in row.values()):
            rows.append(row)

    return rows


@export_bp.route("/import/pdf", methods=["POST"])
@login_required
def import_pdf():
    if not current_user.can_write:
        return jsonify({"error": "Accès refusé (rôle editor ou admin requis)."}), 403

    table = (request.form.get("table") or "contacts").strip()
    mode  = (request.form.get("mode")  or "append").strip()
    if table not in ("contacts", "rocs"):
        return jsonify({"error": "Table invalide."}), 400
    if "file" not in request.files:
        return jsonify({"error": "Aucun fichier reçu."}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Le fichier doit être un PDF (.pdf)."}), 400

    file_bytes = f.read()
    if len(file_bytes) > 10 * 1024 * 1024:
        return jsonify({"error": "Fichier trop volumineux (max 10 Mo)."}), 400

    try:
        pages = _extract_pdf_text(file_bytes)
    except Exception as e:
        return jsonify({"error": f"Impossible de lire le PDF : {str(e)}"}), 400

    if not pages:
        return jsonify({"error": "Aucun texte extractible dans ce PDF (PDF image/scanné non supporté)."}), 400

    # Tenter d'abord la détection tabulaire, puis la détection par blocs (contacts seulement)
    records = _parse_pdf_table(pages, table)
    method  = "tableau"
    if not records and table == "contacts":
        records = _parse_pdf_contacts(pages)
        method  = "blocs"

    if not records:
        return jsonify({
            "error": "Aucune donnée détectable dans ce PDF. "
                     "Le PDF doit contenir du texte sélectionnable (non scanné) "
                     "et être structuré en tableau ou en fiches."
        }), 400

    inserted = 0
    skipped  = 0

    try:
        if mode == "replace":
            if table == "contacts":
                Contact.query.delete(synchronize_session=False)
            else:
                Roc.query.delete(synchronize_session=False)
            db.session.flush()

        for rec in records:
            if table == "contacts":
                if not any([rec.get("societe"), rec.get("nom"), rec.get("email"), rec.get("telephone")]):
                    skipped += 1
                    continue
                db.session.add(Contact(
                    societe    = rec.get("societe",    ""),
                    nom        = rec.get("nom",        ""),
                    prenom     = rec.get("prenom",     ""),
                    fonction   = rec.get("fonction",   ""),
                    email      = rec.get("email",      ""),
                    telephone  = rec.get("telephone",  ""),
                    telephone2 = rec.get("telephone2", ""),
                    notes      = rec.get("notes",      ""),
                    updated_by = current_user.username,
                ))
            else:
                if not any([rec.get("nom_client"), rec.get("roc")]):
                    skipped += 1
                    continue
                db.session.add(Roc(
                    nom_client                = rec.get("nom_client",               ""),
                    roc                       = rec.get("roc",                      ""),
                    trinity                   = rec.get("trinity",                  ""),
                    infogerance               = rec.get("infogerance",              ""),
                    astreinte                 = rec.get("astreinte",                ""),
                    type_contrat              = rec.get("type_contrat",             ""),
                    date_anniversaire_contrat = rec.get("date_anniversaire_contrat",""),
                    updated_by                = current_user.username,
                ))
            inserted += 1

        db.session.add(AuditLog(
            username   = current_user.username,
            action     = "IMPORT_PDF",
            table_name = table,
            detail     = f"Import PDF {table} — {inserted} insérés, {skipped} ignorés (méthode: {method}, mode: {mode})",
            ip_address = request.remote_addr or "",
        ))
        db.session.commit()

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Erreur lors de l'insertion : {str(e)}"}), 500

    return jsonify({
        "ok":       True,
        "inserted": inserted,
        "skipped":  skipped,
        "pages":    len(pages),
        "method":   method,
        "mode":     mode,
        "table":    table,
    })


@export_bp.route("/import/pdf/preview", methods=["POST"])
@login_required
def preview_pdf():
    if not current_user.can_write:
        return jsonify({"error": "Accès refusé."}), 403

    table = (request.form.get("table") or "contacts").strip()
    if "file" not in request.files:
        return jsonify({"error": "Aucun fichier."}), 400

    f = request.files["file"]
    file_bytes = f.read()

    try:
        pages = _extract_pdf_text(file_bytes)
    except Exception as e:
        return jsonify({"error": f"Impossible de lire le PDF : {str(e)}"}), 400

    if not pages:
        return jsonify({"error": "Aucun texte extractible (PDF scanné non supporté)."}), 400

    records = _parse_pdf_table(pages, table)
    method  = "tableau"
    if not records and table == "contacts":
        records = _parse_pdf_contacts(pages)
        method  = "blocs"

    preview = records[:5]

    return jsonify({
        "pages":       len(pages),
        "total_rows":  len(records),
        "method":      method,
        "preview":     preview,
        "fields_found": list({k for r in records for k in r.keys()}),
    })
